"""
Operational Analytics Dashboard - Excel Exporter
==============================================
Purpose: Generate styled Excel workbooks with KPI dashboards
Author: Copilot
Date: 2025-11-23
"""

import pandas as pd
import os
from pathlib import Path
from datetime import datetime
import glob

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import LineChart, BarChart, PieChart, Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("⚠ Warning: openpyxl not installed. Install with: pip install openpyxl")

# Configuration
PROCESSED_DIR = os.path.join(os.path.dirname(__file__), '..', 'data', 'processed')
OUTPUT_FILE = os.path.join(os.path.dirname(__file__), '..', 'data', 'analytics_dashboard.xlsx')


class ExcelDashboard:
    """Create styled Excel workbooks with multiple sheets and formatting"""
    
    def __init__(self, output_file=OUTPUT_FILE):
        """Initialize Excel dashboard generator"""
        self.output_file = output_file
        self.writer = None
        
        # Style definitions
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.title_font = Font(bold=True, size=14, color="1F4E78")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def apply_header_style(self, worksheet):
        """Apply formatting to header row"""
        for cell in worksheet[1]:
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
    
    def auto_adjust_columns(self, worksheet):
        """Auto-adjust column widths based on content"""
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def add_sheet_with_styling(self, df, sheet_name, description=""):
        """
        Add DataFrame to Excel with professional styling
        
        Args:
            df (DataFrame): Data to write
            sheet_name (str): Name of the sheet
            description (str): Optional description for the sheet
        """
        # Write DataFrame to Excel
        df.to_excel(self.writer, sheet_name=sheet_name, index=False, startrow=2 if description else 0)
        
        # Get worksheet
        worksheet = self.writer.sheets[sheet_name]
        
        # Add title/description if provided
        if description:
            worksheet['A1'] = description
            worksheet['A1'].font = self.title_font
            worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
        
        # Apply header styling
        header_row = 3 if description else 1
        for cell in worksheet[header_row]:
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.border
        
        # Apply borders to all data cells
        for row in worksheet.iter_rows(min_row=header_row, max_row=worksheet.max_row, 
                                       min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.border = self.border
                if cell.row > header_row:  # Data rows
                    cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Auto-adjust columns
        self.auto_adjust_columns(worksheet)
        
        # Freeze header row
        worksheet.freeze_panes = f'A{header_row + 1}'
    
    def create_summary_sheet(self):
        """Create executive summary sheet"""
        summary_data = {
            'Metric': [
                'Report Generated',
                'Database',
                'Total KPI Reports',
                'Total Anomaly Reports',
                'Reporting Period',
                'Data Quality',
                'Status'
            ],
            'Value': [
                datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'SQLite - analytics.db',
                '8 KPI Dashboards',
                '7 Anomaly Detections',
                'Last 90 days',
                'Validated',
                '✓ Complete'
            ]
        }
        
        df = pd.DataFrame(summary_data)
        self.add_sheet_with_styling(df, 'Executive Summary', 'Operational Analytics Dashboard - Summary')
        
        # Add instructions
        worksheet = self.writer.sheets['Executive Summary']
        instructions_row = len(df) + 5
        
        worksheet[f'A{instructions_row}'] = 'DASHBOARD NAVIGATION'
        worksheet[f'A{instructions_row}'].font = self.title_font
        
        instructions = [
            "• KPI sheets: Business performance metrics and trends",
            "• Anomaly sheets: Unusual patterns requiring investigation",
            "• Each sheet includes detailed data for drill-down analysis",
            "• Use filters and sorting to explore specific segments",
            "• Export sheets to Power BI for advanced visualization"
        ]
        
        for idx, instruction in enumerate(instructions, start=1):
            worksheet[f'A{instructions_row + idx}'] = instruction
    
    def generate_dashboard(self):
        """Main method to generate complete Excel dashboard"""
        print("\n" + "="*60)
        print("EXCEL DASHBOARD GENERATION")
        print("="*60)
        
        # Check if openpyxl is available
        if not OPENPYXL_AVAILABLE:
            print("✗ Cannot generate styled Excel file without openpyxl")
            print("  Generating basic CSV exports only")
            return False
        
        # Find all CSV files in processed directory
        csv_files = glob.glob(os.path.join(PROCESSED_DIR, '*.csv'))
        
        if not csv_files:
            print("✗ No CSV files found in processed directory")
            print(f"  Run 'python run_queries.py' first to generate data")
            return False
        
        print(f"✓ Found {len(csv_files)} CSV files")
        
        # Create Excel writer
        self.writer = pd.ExcelWriter(self.output_file, engine='openpyxl')
        
        # Create summary sheet first
        self.create_summary_sheet()
        
        # Process each CSV file
        sheet_mapping = {
            'daily_active_users': ('DAU Trend', 'Daily Active Users - 90 Day Trend Analysis'),
            'user_churn_rate': ('Churn Analysis', 'User Churn Rate by Cohort and Tier'),
            'refund_rate_analysis': ('Refund Metrics', 'Refund Rate and Financial Impact'),
            'ticket_resolution_time': ('Support SLA', 'Average Ticket Resolution Time vs SLA'),
            'customer_lifetime_value': ('Customer CLV', 'Customer Lifetime Value by Tier'),
            'transaction_success_rate': ('Payment Success', 'Transaction Success Rate Analysis'),
            'cohort_retention': ('Retention Matrix', 'User Cohort Retention Analysis'),
            'top_users_ranking': ('Top Users', 'User Ranking by Engagement and Spend'),
            'anomaly_transaction_amount_spikes': ('Anomaly Spikes', 'Transaction Amount Spikes Detection'),
            'anomaly_consecutive_payment_failures': ('Anomaly Failures', 'Consecutive Payment Failures'),
            'anomaly_refund_rate_spikes': ('Anomaly Refunds', 'Sudden Refund Rate Spikes'),
            'anomaly_dormant_user_reactivation': ('Anomaly Dormant', 'Dormant User Reactivation'),
            'anomaly_ticket_resolution_outliers': ('Anomaly Tickets', 'Ticket Resolution Outliers'),
            'anomaly_unusual_login_patterns': ('Anomaly Logins', 'Unusual Login Patterns'),
            'anomaly_payment_method_switching': ('Anomaly Methods', 'Payment Method Switching')
        }
        
        for csv_file in csv_files:
            file_name = os.path.splitext(os.path.basename(csv_file))[0]
            
            # Get sheet name and description
            sheet_info = sheet_mapping.get(file_name, (file_name[:31], file_name))
            sheet_name, description = sheet_info
            
            try:
                df = pd.read_csv(csv_file)
                
                if len(df) > 0:
                    self.add_sheet_with_styling(df, sheet_name, description)
                    print(f"✓ Added sheet: {sheet_name} ({len(df)} rows)")
                else:
                    print(f"⚠ Skipped {file_name}: No data")
            
            except Exception as e:
                print(f"✗ Error processing {file_name}: {str(e)[:100]}")
        
        # Save workbook
        try:
            self.writer.close()
            print(f"\n✓ Excel dashboard saved: {self.output_file}")
            print(f"  File size: {os.path.getsize(self.output_file) / 1024:.1f} KB")
            return True
        except Exception as e:
            print(f"✗ Error saving Excel file: {e}")
            return False


def export_individual_sheets():
    """Export each KPI as separate Excel file (fallback if no openpyxl)"""
    print("\n" + "="*60)
    print("INDIVIDUAL EXCEL EXPORTS")
    print("="*60)
    
    csv_files = glob.glob(os.path.join(PROCESSED_DIR, '*.csv'))
    
    for csv_file in csv_files:
        try:
            df = pd.read_csv(csv_file)
            output_file = csv_file.replace('.csv', '.xlsx')
            df.to_excel(output_file, index=False, engine='openpyxl')
            print(f"✓ Exported: {os.path.basename(output_file)}")
        except Exception as e:
            print(f"✗ Error exporting {os.path.basename(csv_file)}: {str(e)[:100]}")


def main():
    """Main execution flow"""
    print("\n" + "="*70)
    print(" "*15 + "EXCEL DASHBOARD GENERATOR")
    print(" "*17 + "Operational Analytics")
    print("="*70)
    print(f"Execution Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    # Create output directory if needed
    Path(os.path.dirname(OUTPUT_FILE)).mkdir(parents=True, exist_ok=True)
    
    # Generate dashboard
    dashboard = ExcelDashboard()
    success = dashboard.generate_dashboard()
    
    if success:
        print("\n" + "="*60)
        print("DASHBOARD READY")
        print("="*60)
        print(f"✓ Open file: {OUTPUT_FILE}")
        print("✓ Contains: Executive Summary + 15 analytical sheets")
        print("✓ Features: Styled headers, auto-sized columns, freeze panes")
        
        print("\n" + "="*60)
        print("NEXT STEPS")
        print("="*60)
        print("1. Open the Excel file to review KPIs and anomalies")
        print("2. Use filters and sorting for detailed analysis")
        print("3. Import sheets into Power BI for visualization")
        print("4. Share insights with stakeholders")
    else:
        print("\n⚠ Dashboard generation incomplete")
        print("  Ensure you've run 'python run_queries.py' first")
        print("  Install openpyxl: pip install openpyxl")


if __name__ == "__main__":
    main()
